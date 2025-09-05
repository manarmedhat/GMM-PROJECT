import pandas as pd
from sklearn.mixture import GaussianMixture
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from copy import copy

# Step 1: Read the description line
with open('Female data.csv', 'r', encoding='utf-8') as f:
    description_line = f.readline().strip()

# Step 2: Read the dataset
data = pd.read_csv('Female data.csv', skiprows=1)

# Drop first column if it is string
if data.dtypes[0] == 'object':
    data = data.drop(data.columns[0], axis=1)

# Fix decimals and clean
data = data.replace('٫', '.', regex=True)
data = data.apply(pd.to_numeric, errors='coerce')
data = data.dropna()

# Step 3: Standardize and cluster
scaler = StandardScaler()
X_scaled = scaler.fit_transform(data)

gmm = GaussianMixture(n_components=4, random_state=42)
clusters = gmm.fit_predict(X_scaled)
data['cluster'] = clusters

# Step 4: Build summaries
summary_rows = []
ranges_rows = []
for c in range(4):
    cluster_data = data[data['cluster'] == c]
    means = cluster_data.mean(numeric_only=True).round(0)
    count = len(cluster_data)
    
    summary = {'Cluster': c, 'Count': count}
    for feat, val in means.items():
        if feat != 'cluster':
            summary[f'{feat}_Mean'] = val
    summary_rows.append(summary)
    
    min_vals = cluster_data.min(numeric_only=True)
    max_vals = cluster_data.max(numeric_only=True)
    for feat in min_vals.index:
        if feat != 'cluster':
            ranges_rows.append({
                'Cluster': c,
                'Feature': feat,
                'Min': round(min_vals[feat], 0),
                'Max': round(max_vals[feat], 0)
            })

summary_df = pd.DataFrame(summary_rows)
ranges_df = pd.DataFrame(ranges_rows)
counts_df = data['cluster'].value_counts().reset_index()
counts_df.columns = ['Cluster', 'Count']

# Step 5: Write results into a new Excel (without formatting)
output_file = "GMM_results_new_female.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    pd.DataFrame([{"Description": description_line}]).to_excel(writer, sheet_name='Description', index=False)
    summary_df.to_excel(writer, sheet_name='Cluster_Summary', index=False)
    ranges_df.to_excel(writer, sheet_name='Cluster_Ranges', index=False)
    counts_df.to_excel(writer, sheet_name='Cluster_Counts', index=False)
    data.to_excel(writer, sheet_name='Data_with_Clusters', index=False)

# Step 6: Copy formatting, column widths, and row heights from old file
source_wb = load_workbook("GMM_results.xlsx")
target_wb = load_workbook(output_file)

for sheet_name in source_wb.sheetnames:
    if sheet_name in target_wb.sheetnames:
        source_ws = source_wb[sheet_name]
        target_ws = target_wb[sheet_name]
        
        # Copy cell formatting
        for row in source_ws.iter_rows():
            for cell in row:
                tgt_cell = target_ws[cell.coordinate]
                if cell.has_style:
                    tgt_cell.font = copy(cell.font)
                    tgt_cell.border = copy(cell.border)
                    tgt_cell.fill = copy(cell.fill)
                    tgt_cell.number_format = copy(cell.number_format)
                    tgt_cell.protection = copy(cell.protection)
                    tgt_cell.alignment = copy(cell.alignment)
        
        # Copy column widths
        for col_letter, col_dim in source_ws.column_dimensions.items():
            if col_letter in target_ws.column_dimensions:
                target_ws.column_dimensions[col_letter].width = col_dim.width

        # Copy row heights
        for row_idx, row_dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[row_idx].height = row_dim.height

target_wb.save("GMM_results_final_female.xlsx")

